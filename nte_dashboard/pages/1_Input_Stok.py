"""
Page: Input Stok — st.data_editor (mirip Google Sheets)
Semua warehouse dalam 1 tabel, edit langsung, simpan sekali klik.
"""

import streamlit as st
import pandas as pd
import io
import sys, os
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from utils.database import init_db, upsert_stok, get_stok, delete_stok
from data.master_data import (
    AREA_CONFIG, ALL_OPERATORS, NTE_STATUS,
    NTE_CATALOG_BY_OPERATOR, get_type_to_jenis
)

init_db()

st.set_page_config(
    page_title="Input Stok | NTE Dashboard",
    page_icon="✏️",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Space+Grotesk:wght@600;700&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif}
.page-header{background:linear-gradient(135deg,#1E3A5F,#2E6DA4);color:white;
  padding:1.25rem 2rem;border-radius:12px;margin-bottom:1rem}
.page-header h2{font-family:'Space Grotesk',sans-serif;margin:0;font-size:1.4rem}
.page-header p{margin:.2rem 0 0;opacity:.8;font-size:.85rem}
.info-bar{background:#EBF5FF;border:1px solid #BDE0FF;border-radius:8px;
  padding:.65rem 1rem;font-size:.83rem;color:#0D47A1;margin-bottom:.75rem}
.stat-row{display:flex;gap:.75rem;margin:.75rem 0;flex-wrap:wrap}
.stat-box{background:white;border:1px solid #E0EAF5;border-radius:8px;
  padding:.5rem 1rem;font-size:.82rem;color:#555;min-width:110px}
.stat-box b{display:block;font-size:1.2rem;color:#1E3A5F;font-family:'Space Grotesk',sans-serif}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def short_wh(wh: str) -> str:
    """Singkat nama WH untuk header kolom tabel."""
    return (wh
        .replace("TA SO INV ", "")
        .replace("TA SO CCAN ", "")
        .replace("TA SO TIF ", "")
        .replace(" WH", "")
        .strip()
    )

def build_template(catalog: dict, whs: list, status_list: list) -> pd.DataFrame:
    """Buat DataFrame kosong berisi semua kombinasi jenis × type × status × WH."""
    rows = []
    for jenis, types in catalog.items():
        for type_nte in types:
            for status in status_list:
                row = {"JENIS 2": jenis, "TYPE": type_nte, "STATUS": status}
                for wh in whs:
                    row[wh] = 0
                rows.append(row)
    return pd.DataFrame(rows)

def fill_from_db(df_tpl: pd.DataFrame, df_db: pd.DataFrame, whs: list) -> pd.DataFrame:
    """Isi template dengan data dari database."""
    df = df_tpl.copy()
    if df_db.empty:
        return df
    for _, row in df_db.iterrows():
        wh  = row["warehouse"]
        if wh not in df.columns:
            continue
        mask = (df["TYPE"] == row["type_nte"]) & (df["STATUS"] == row["status_nte"])
        if mask.any():
            df.loc[mask, wh] = int(row["closing_stock"])
    return df

def make_short_map(whs: list) -> dict:
    """Buat mapping nama panjang → nama pendek, hindari duplikat."""
    mapping = {}
    seen    = {}
    for wh in whs:
        sh = short_wh(wh)
        if sh in seen.values():
            sh = sh + " 2"
        mapping[wh] = sh
        seen[wh]    = sh
    return mapping

def export_to_excel(df: pd.DataFrame, wh_short_cols: list,
                    operator: str, area: str, tanggal: str) -> bytes:
    """Export tabel ke Excel."""
    wb  = Workbook()
    ws  = wb.active
    ws.title = f"{operator} {area}"[:28]
    cols_out = ["JENIS 2", "TYPE", "STATUS"] + wh_short_cols + ["GRAND TOTAL"]
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Arial")
    hdr_al   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(cols_out, 1):
        c = ws.cell(1, ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = hdr_al
        ws.column_dimensions[get_column_letter(ci)].width = (
            18 if ci == 1 else 32 if ci == 2 else 12 if ci == 3 else 10
        )
    ws.row_dimensions[1].height = 32
    safe_cols = [c for c in cols_out if c in df.columns]
    for ri, row in enumerate(df[safe_cols].itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, value=val)
        ws.row_dimensions[ri].height = 18
    ws.freeze_panes = "D2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("### ⚙️ Pilih Laporan")
    sel_date = st.date_input("📅 Tanggal", value=date.today())
    sel_op   = st.selectbox("🏢 Operator", ALL_OPERATORS)
    op_keys  = [k for k, v in AREA_CONFIG.items() if v["operator"] == sel_op]
    sel_key  = st.selectbox("📍 Area", op_keys)
    cfg      = AREA_CONFIG[sel_key]
    whs      = cfg["warehouses"]
    catalog  = NTE_CATALOG_BY_OPERATOR.get(sel_op, {})
    t2j      = get_type_to_jenis(sel_op)

    st.divider()
    st.caption(f"**{len(whs)} warehouse** terdaftar")
    st.caption(f"**{sum(len(t) for t in catalog.values())} type NTE** di katalog")
    st.divider()

    st.markdown("**Navigasi cepat**")
    c1, c2 = st.columns(2)
    if c1.button("◀ Kemarin", use_container_width=True):
        st.session_state["sel_date_override"] = str(date.today() - timedelta(days=1))
    if c2.button("Hari ini", use_container_width=True):
        st.session_state["sel_date_override"] = str(date.today())

# ══════════════════════════════════════════════════════════════════════════════
# HEADER
# ══════════════════════════════════════════════════════════════════════════════

st.markdown(f"""
<div class="page-header">
    <h2>✏️ Input Stok — {sel_op} {cfg['area']}</h2>
    <p>Edit langsung di tabel · semua warehouse dalam satu layar · mirip Google Sheets</p>
</div>
""", unsafe_allow_html=True)

st.markdown(f"""
<div class="info-bar">
    📅 <b>{sel_date.strftime('%A, %d %B %Y')}</b>
    &nbsp;·&nbsp; 🏢 <b>{sel_op}</b>
    &nbsp;·&nbsp; 📍 <b>{cfg['area']}</b>
    &nbsp;·&nbsp; 🏭 <b>{len(whs)} WH</b>
    &nbsp;&nbsp;|&nbsp;&nbsp;
    💡 Klik sel angka untuk edit · Tab pindah kolom · Enter konfirmasi
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# BUILD TABLE
# ══════════════════════════════════════════════════════════════════════════════

if not catalog:
    st.warning("⚠️ Katalog NTE untuk operator ini belum diisi. Edit `data/master_data.py`.")
    st.stop()

# Load data existing
df_db  = get_stok(str(sel_date), operator=sel_op, area_key=sel_key)

# Bangun & isi template
df_tpl = build_template(catalog, whs, NTE_STATUS)
df_tpl = fill_from_db(df_tpl, df_db, whs)

# Short name map
short_map     = make_short_map(whs)
rev_map       = {v: k for k, v in short_map.items()}
short_wh_cols = [short_map[w] for w in whs]

# Rename kolom untuk tampilan
df_display = df_tpl.rename(columns=short_map)

# ── Statistik ──────────────────────────────────────────────────────────────
total_unit  = int(df_tpl[whs].sum().sum())
wh_filled   = sum(1 for w in whs if df_tpl[w].sum() > 0)
rows_filled = int((df_tpl[whs].sum(axis=1) > 0).sum())

st.markdown(f"""
<div class="stat-row">
  <div class="stat-box"><b>{total_unit:,}</b>Total unit</div>
  <div class="stat-box"><b>{wh_filled}/{len(whs)}</b>WH ada data</div>
  <div class="stat-box"><b>{rows_filled}</b>Baris terisi</div>
  <div class="stat-box"><b>{len(catalog)}</b>Jenis NTE</div>
</div>
""", unsafe_allow_html=True)

# ── Filter ─────────────────────────────────────────────────────────────────
with st.expander("🔍 Filter tampilan", expanded=False):
    fc1, fc2 = st.columns(2)
    filter_jenis  = fc1.multiselect(
        "Jenis NTE", list(catalog.keys()), default=list(catalog.keys())
    )
    filter_status = fc2.multiselect(
        "Status", NTE_STATUS, default=NTE_STATUS
    )
    show_zeros = st.checkbox("Tampilkan baris stok = 0", value=True)

df_view = df_display[
    df_display["JENIS 2"].isin(filter_jenis) &
    df_display["STATUS"].isin(filter_status)
].copy()

if not show_zeros:
    df_view = df_view[(df_view[short_wh_cols].sum(axis=1) > 0)]

# Tambah GRAND TOTAL read-only
df_view["GRAND TOTAL"] = df_view[short_wh_cols].sum(axis=1)

# ── Column config ──────────────────────────────────────────────────────────
col_config = {
    "JENIS 2":     st.column_config.TextColumn("JENIS 2",   width="medium",  disabled=True),
    "TYPE":        st.column_config.TextColumn("TYPE",      width="large",   disabled=True),
    "STATUS":      st.column_config.TextColumn("STATUS",    width="small",   disabled=True),
    "GRAND TOTAL": st.column_config.NumberColumn("GRAND TOTAL", format="%d", width="small", disabled=True),
}
for wh_short in short_wh_cols:
    col_config[wh_short] = st.column_config.NumberColumn(
        label=wh_short,
        min_value=0,
        max_value=99999,
        step=1,
        format="%d",
        width="small",
    )

# ══════════════════════════════════════════════════════════════════════════════
# DATA EDITOR  ← ini pengganti form + input manual
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("#### 📊 Tabel Input Stok")
st.caption(
    f"Menampilkan **{len(df_view)} baris**. "
    "Kolom JENIS, TYPE, STATUS, GRAND TOTAL terkunci — hanya kolom WH yang bisa diedit."
)

edited = st.data_editor(
    df_view,
    column_config=col_config,
    use_container_width=True,
    hide_index=True,
    num_rows="fixed",
    height=min(720, 44 + len(df_view) * 35),
    key=f"tbl_{sel_key}_{sel_date}",
)

# Update GRAND TOTAL di edited (recalculate)
if edited is not None and not edited.empty:
    edited["GRAND TOTAL"] = edited[short_wh_cols].sum(axis=1)

# ── Subtotal per WH ────────────────────────────────────────────────────────
if edited is not None and not edited.empty:
    st.markdown("**Subtotal per Warehouse:**")
    sub_cols = st.columns(len(short_wh_cols) + 1)
    g_total  = 0
    for i, sh in enumerate(short_wh_cols):
        val = int(edited[sh].sum()) if sh in edited.columns else 0
        sub_cols[i].metric(sh, f"{val:,}")
        g_total += val
    sub_cols[-1].metric("**Grand Total**", f"**{g_total:,}**")

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# ACTIONS
# ══════════════════════════════════════════════════════════════════════════════

ca, cb, cc = st.columns([3, 2, 2])

# ── Simpan ─────────────────────────────────────────────────────────────────
with ca:
    if st.button("💾 Simpan Semua ke Database", type="primary", use_container_width=True):
        if edited is None or edited.empty:
            st.warning("Tidak ada data.")
        else:
            saved   = 0
            with st.spinner("Menyimpan..."):
                for _, row in edited.iterrows():
                    type_nte   = row["TYPE"]
                    status_nte = row["STATUS"]
                    jenis_nte  = t2j.get(type_nte, row["JENIS 2"])
                    for sh in short_wh_cols:
                        wh_orig = rev_map.get(sh, sh)
                        try:   val = int(row[sh])
                        except: val = 0
                        if val > 0:
                            upsert_stok(
                                tanggal=str(sel_date), operator=sel_op,
                                area=cfg["area"], area_key=sel_key,
                                warehouse=wh_orig, jenis_nte=jenis_nte,
                                type_nte=type_nte, status_nte=status_nte,
                                closing_stock=val,
                            )
                            saved += 1
            st.success(f"✅ **{saved} entri** berhasil disimpan!")
            st.balloons()
            st.rerun()

# ── Hapus ──────────────────────────────────────────────────────────────────
with cb:
    if st.button("🗑️ Hapus Data Tanggal Ini", use_container_width=True):
        st.session_state["confirm_delete"] = True

if st.session_state.get("confirm_delete"):
    st.warning(
        f"⚠️ Hapus semua data **{sel_op} {cfg['area']}** untuk "
        f"**{sel_date}**?"
    )
    cd1, cd2 = st.columns(2)
    if cd1.button("✅ Ya, hapus", type="secondary", use_container_width=True):
        for wh in whs:
            delete_stok(str(sel_date), sel_op, wh)
        st.session_state["confirm_delete"] = False
        st.success("Data dihapus.")
        st.rerun()
    if cd2.button("❌ Batal", use_container_width=True):
        st.session_state["confirm_delete"] = False
        st.rerun()

# ── Export ──────────────────────────────────────────────────────────────────
with cc:
    if edited is not None and not edited.empty:
        xlsx = export_to_excel(
            edited, short_wh_cols,
            sel_op, cfg["area"], str(sel_date)
        )
        st.download_button(
            label="⬇️ Export tabel ini ke Excel",
            data=xlsx,
            file_name=f"input_{sel_op}_{cfg['area']}_{sel_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

# ── Tips ───────────────────────────────────────────────────────────────────
with st.expander("💡 Tips penggunaan", expanded=False):
    st.markdown("""
    **Cara edit tabel (mirip Google Sheets):**
    - **Klik** sel angka pada kolom WH → ketik nilai → **Enter**
    - **Tab** untuk pindah ke kolom berikutnya
    - **Shift+Tab** untuk kembali ke kolom sebelumnya
    - **Scroll horizontal** jika WH tidak muat di layar

    **Alur kerja harian yang disarankan:**
    1. Pilih Operator & Area di sidebar
    2. Isi angka stok closing untuk setiap WH
    3. Cek Subtotal per WH di bawah tabel
    4. Klik **Simpan Semua ke Database**
    5. Dashboard Laporan Harian otomatis terupdate

    **Catatan:**
    - Baris dengan nilai 0 **tidak disimpan** (diabaikan)
    - Data lama untuk tanggal yang sama akan **ditimpa** jika diubah
    - Gunakan filter Jenis NTE jika hanya ingin isi sebagian
    """)
