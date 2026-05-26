"""Export ke Excel — format STOCK NTE (sesuai G-Sheet)"""

import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

NAVY    = "1E3A5F"; BLUE = "2E6DA4"; LBLUE = "D9EAF7"
RED     = "C0392B"; REDBG = "FADBD8"
GREENBG = "D5F5E3"; GREENTXT = "1E8449"
YELBG   = "FFF3CD"; YELTXT   = "856404"
GRAYBG  = "F2F2F2"; WHITE    = "FFFFFF"; MIDGRAY = "CCCCCC"

def _side(c=MIDGRAY): return Side(style="thin", color=c)
def _bdr(c=MIDGRAY):  b = _side(c); return Border(left=b,right=b,top=b,bottom=b)
def _hdr(cell, bg=NAVY, fg=WHITE, bold=True, sz=10, align="center"):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(color=fg, bold=bold, size=sz, name="Arial")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border    = _bdr()
def _cell(cell, align="center", bg=None, bold=False, color="222222", sz=10):
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border    = _bdr()
    cell.font      = Font(bold=bold, color=color, size=sz, name="Arial")
    if bg: cell.fill = PatternFill("solid", fgColor=bg)


def export_rekap_area_excel(pivot_df, detail_df, area, tanggal, warehouses):
    wb = Workbook()
    ws = wb.active
    ws.title = area[:28]

    n  = 3 + len(warehouses) + 1
    lc = get_column_letter(n)

    # Title rows
    ws.merge_cells(f"A1:{lc}1")
    ws["A1"] = f"STOCK NTE {area.upper()}"
    ws["A1"].font      = Font(bold=True, size=14, color=NAVY, name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill      = PatternFill("solid", fgColor=LBLUE)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{lc}2")
    ws["A2"] = f"Tanggal: {tanggal}"
    ws["A2"].font      = Font(italic=True, size=10, color="555555", name="Arial")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 15

    # Sub-header (COUNTA of SN | WH SO SESUAI SCMT)
    ws.merge_cells("A3:C3"); ws["A3"] = "COUNTA of SN"
    _hdr(ws["A3"], bg=GRAYBG, fg=NAVY, sz=9)
    ws.merge_cells(f"D3:{lc}3"); ws.cell(3,4).value = "WH SO (SESUAI SCMT)"
    _hdr(ws.cell(3,4), bg=GRAYBG, fg=NAVY, sz=9)
    ws.row_dimensions[3].height = 15

    # Column headers
    hdrs = ["JENIS 2","STATUS","TYPE"] + warehouses + ["Grand Total"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(4, ci, value=h)
        if   ci <= 3:    _hdr(c, bg=NAVY,  sz=9)
        elif ci == n:    _hdr(c, bg=RED,   sz=9)
        else:            _hdr(c, bg=BLUE,  sz=9)
        ws.column_dimensions[get_column_letter(ci)].width = (
            18 if ci==1 else 12 if ci==2 else 34 if ci==3 else 10
        )
    ws.row_dimensions[4].height = 36

    # Data rows
    RS = 5; ri = RS; prev_j = None
    for _, r in pivot_df.iterrows():
        jenis  = r.get("jenis_nte","");  type_ = r.get("type_nte","")
        status = r.get("status_nte",""); grand = int(r.get("Grand Total", 0))

        jc = ws.cell(ri, 1, value=(jenis if jenis!=prev_j else ""))
        _cell(jc, align="left", bg="EAF3FB", bold=bool(jenis!=prev_j), color=NAVY, sz=9)

        sc = ws.cell(ri, 2, value=status)
        _cell(sc, bg=(GREENBG if status=="NTE BARU" else YELBG),
              color=(GREENTXT if status=="NTE BARU" else YELTXT), bold=True, sz=9)

        tc = ws.cell(ri, 3, value=type_)
        _cell(tc, align="left", sz=9)

        for wi, wh in enumerate(warehouses, 4):
            val = int(r.get(wh, 0) or 0)
            c   = ws.cell(ri, wi, value=(val if val>0 else None))
            # Heat map coloring
            if val == 0:    _cell(c, color=MIDGRAY, sz=9)
            else:           _cell(c, bg="F0FBF0", color="1A6632", sz=9)

        gtc = ws.cell(ri, n, value=(grand if grand>0 else None))
        _cell(gtc, bg=REDBG, color=RED, bold=True, sz=10)

        prev_j = jenis; ri += 1

    # Grand total row
    ws.merge_cells(f"A{ri}:C{ri}")
    _hdr(ws.cell(ri,1, value="Grand Total"), bg=NAVY, sz=10)
    for wi in range(4, n):
        cl = get_column_letter(wi)
        c  = ws.cell(ri, wi, value=f"=SUM({cl}{RS}:{cl}{ri-1})")
        _hdr(c, bg=BLUE, sz=9)
    gtcl = get_column_letter(n)
    _hdr(ws.cell(ri, n, value=f"=SUM({gtcl}{RS}:{gtcl}{ri-1})"), bg=RED, sz=11)
    ws.row_dimensions[ri].height = 20
    ws.freeze_panes = "D5"

    # Sheet 2: Detail
    if not detail_df.empty:
        ws2 = wb.create_sheet("Data Detail")
        for r2i, row in enumerate(dataframe_to_rows(detail_df, index=False, header=True), 1):
            for c2i, val in enumerate(row, 1):
                cell = ws2.cell(r2i, c2i, value=val)
                if r2i==1: _hdr(cell, sz=9)
                else:      _cell(cell, align="left", sz=9)
        for ci in range(1, len(detail_df.columns)+1):
            ws2.column_dimensions[get_column_letter(ci)].width = 22

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def build_excel_template(area_keys: list) -> bytes:
    from data.master_data import AREA_CONFIG, NTE_CATALOG_BY_OPERATOR, NTE_STATUS
    wb = Workbook()
    ws = wb.active; ws.title = "Template Input Stok"
    hdrs = ["tanggal","operator","area","area_key","warehouse",
            "jenis_nte","type_nte","status_nte","closing_stock"]
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws.cell(1,ci,value=h))
        ws.column_dimensions[get_column_letter(ci)].width = 26

    examples = [
        ["2025-05-19","TELKOMSEL","BANDUNG","TELKOMSEL - BANDUNG",
         "TA SO INV AHMAD YANI WH","ONT DUAL BAND","ONT_FIBERHOME_HG6145D2","NTE BARU",247],
        ["2025-05-19","TELKOM","BANDUNG","TELKOM - BANDUNG",
         "TA SO CCAN AHMAD YANI WH","ONT SINGLE BAND","ONT_FIBERHOME_AN5506-04-FS","NTE BARU",10],
        ["2025-05-19","TIF","SOREANG","TIF - SOREANG",
         "TA SO TIF KADIPATEN WH","ONT PREMIUM","ONT_FIBERHOME_HG6145F1","REFURBISH",5],
    ]
    for ri, ex in enumerate(examples, 2):
        for ci, val in enumerate(ex, 1):
            c = ws.cell(ri,ci,value=val)
            c.fill = PatternFill("solid",fgColor="F0F7FF")
            _cell(c, align="left", sz=9)

    ws2 = wb.create_sheet("Referensi")
    ref_h = ["operator","area","area_key","warehouse","jenis_nte","type_nte","status_nte"]
    for ci, h in enumerate(ref_h,1):
        _hdr(ws2.cell(1,ci,value=h))
        ws2.column_dimensions[get_column_letter(ci)].width = 30

    ri2 = 2
    for ak, cfg in AREA_CONFIG.items():
        op  = cfg["operator"]
        cat = NTE_CATALOG_BY_OPERATOR.get(op, {})
        for wh in cfg["warehouses"]:
            for jenis, types in cat.items():
                for t in types:
                    for st in NTE_STATUS:
                        for ci2, val in enumerate(
                            [op, cfg["area"], ak, wh, jenis, t, st], 1
                        ):
                            _cell(ws2.cell(ri2,ci2,value=val), align="left", sz=9)
                        ri2 += 1
    ws2.freeze_panes = "A2"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()
